import openpyxl
from openpyxl.utils import range_boundaries

def get_real_max_row(sheet):
    for row in range(sheet.max_row, 0, -1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value is not None:
                return row
    return 1

def get_header_value(sheet, row, col, merged_cells):
    for merged_range in merged_cells:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_row <= row <= max_row and min_col <= col <= max_col:
            return sheet.cell(row=min_row, column=min_col).value
    return sheet.cell(row=row, column=col).value

def excel_to_html(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = wb.active
    
    real_max_row = get_real_max_row(sheet)
    max_cols = sheet.max_column
    merged_cells = sheet.merged_cells.ranges
    
    html = ['<table id="data-spreadsheet" class="display" style="width:100%">']
    html.append('  <thead>')
    html.append('    <tr class="dt-layout-row">')
    html.append('      <th>SectionTracker</th>')
    
    for col_idx in range(1, max_cols + 1):
        r1_val = get_header_value(sheet, 1, col_idx, merged_cells)
        r2_val = get_header_value(sheet, 2, col_idx, merged_cells)
        
        if r1_val == r2_val:
            final_header = r1_val if r1_val else f"Column {col_idx}"
        else:
            parent = f"{r1_val} > " if r1_val else ""
            parent = ""
            child = r2_val if r2_val else ""
            final_header = f"{parent}{child}"
            
        # --- FIX: Set static widths on the master definition row ---
        width_style = ""
        if col_idx == 1:   # Method Column
            width_style = ' style="width: 20%;"'
        elif col_idx == 2:   # Attribute Column
            width_style = ' style="width: 15%;"'
        elif col_idx == 3:   # Description Column
            width_style = ' style="width: 12%;"'
        elif col_idx == 4: # Updating in Running state Column
            width_style = ' style="width: 12%;"'
        elif col_idx == 5: # Updating in POWEROFF state Column
            width_style = ' style="width: 12%;"'
        elif col_idx == 6: # API Column
            width_style = ' style="width: 5%;"'
        elif col_idx == 7: # CLI Column
            width_style = ' style="width: 5%;"'
        elif col_idx == 8: # API Column
            width_style = ' style="width: 5%;"'
        elif col_idx == 9: # API Column
            width_style = ' style="width: 20%;"'
            
        html.append(f'      <th{width_style}>{final_header}</th>')
        # ------------------------------------------------------------
        
    html.append('    </tr>')
    html.append('  </thead>\n  <tbody>')
    
    current_section = "General"
    
    for row_idx in range(1, real_max_row + 1):
        if row_idx <= 2:
            continue
            
        is_section_heading = False
        for merged_range in merged_cells:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            if min_row == row_idx and min_col == 1 and (max_col - min_col + 1) == max_cols:
                is_section_heading = True
                break
        
        if is_section_heading:
            current_section = sheet.cell(row=row_idx, column=1).value or "General"
            continue 

        html.append('    <tr>')
        html.append(f'      <td>{current_section}</td>')
        
        for col_idx in range(1, max_cols + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            val = cell.value if cell.value is not None else ""
            
            if col_idx == 2 and val:
                html.append(f'      <td class="truncated-attribute" title="{val}">{val}</td>')
            else:
                html.append(f'      <td>{val}</td>')
            
        html.append('    </tr>')
        
    html.append('  </tbody>\n</table>')
    return '\n'.join(html)

# --- FIX: Inject table precisely into the Marker location ---
def inject_table_into_markdown(md_file_path, xlsx_path, marker):
    # 1. Generate the HTML table block with its Hugo shortcode packaging
    html_table = excel_to_html(xlsx_path)
    shortcode_payload = f"{{{{< vm-methods-table >}}}}\n{html_table}\n{{{{< /vm-methods-table >}}}}"
    
    # 2. Read the current contents written by your team members
    try:
        with open(md_file_path, 'r', encoding='utf-8') as file:
            md_content = file.read()
    except FileNotFoundError:
        print(f"Error: Could not find {md_file_path}. Make sure the file exists with the placeholder comment.")
        return

    # If the file already has a built table from a previous run, find the markers and clean it out
    if f"{marker}\n{{{{< vm-methods-table >}}}}" in md_content:
        # Splits the file at the historical payload run and preserves the original template frame
        parts = md_content.split(f"{marker}\n{{{{< vm-methods-table >}}}}")
        header_part = parts[0]
        footer_part = parts[1].split(f"{{{{< /vm-methods-table >}}}}")[-1]
        md_content = f"{header_part}{marker}{footer_part}"

    # 3. Inject our fresh payload right below the pristine marker anchor
    if marker in md_content:
        updated_content = md_content.replace(marker, f"{marker}\n{shortcode_payload}")
        
        # 4. Save it back to the Hugo repository path
        with open(md_file_path, 'w', encoding='utf-8') as file:
            file.write(updated_content)
        print("Success: Table successfully injected into the Markdown placeholder!")
    else:
        print(f"Warning: Marker '{marker}' was not found inside the markdown file. No injection occurred.")

html_table = excel_to_html('assets/tables/vm_update_methods.xlsx')
markdown_output = f"{{{{< vm-methods-table >}}}}\n{html_table}\n{{{{< /vm-methods-table >}}}}"

MD_FILE = "content/product/operation_references/configuration_references/vm_update_methods.md"
XLSX_PATH = "assets/tables/vm_update_methods.xlsx"
inject_table_into_markdown(MD_FILE, XLSX_PATH, "<!-- VM METHODS TABLE -->")